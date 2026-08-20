from __future__ import annotations

import logging
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from models import Localization

logger = logging.getLogger(__name__)

MAX_LANGUAGES = 8

LANGUAGE_COLUMNS = ["Default Language"] + [f"Additional Language {i}" for i in range(1, 8)]
TITLE_COLUMNS = ["Title"] + [f"Additional Title {i}" for i in range(1, 8)]
BODY_COLUMNS = ["Body"] + [f"Additional Body {i}" for i in range(1, 8)]
LINK_COLUMNS = ["Link"] + [f"Additional Link {i}" for i in range(1, 8)]
DISPLAY_LINK_COLUMNS = ["Display Link"] + [f"Additional Display Link {i}" for i in range(1, 8)]

REQUIRED_COLUMNS = LANGUAGE_COLUMNS + TITLE_COLUMNS + BODY_COLUMNS + LINK_COLUMNS
SPECIAL_AD_CATEGORY_COUNTRY = "Special Ad Category Country"
ALLOWED_COLUMNS = set(
    REQUIRED_COLUMNS + DISPLAY_LINK_COLUMNS + [SPECIAL_AD_CATEGORY_COUNTRY]
)


class ExcelTemplateError(RuntimeError):
    pass


@dataclass
class TemplateInfo:
    path: Path
    workbook: Workbook
    worksheet: Worksheet
    headers: dict[str, int]
    header_names: list[str]
    max_row: int
    max_column: int
    dimensions: str | None
    data_rows: list[int]


@dataclass
class RowSource:
    row: int
    title: str
    body: str
    link: str
    geo: str | None
    product: str = ""


def backup_workbook(src: Path) -> Path:
    dest = src.with_name(f"{src.stem}_backup{src.suffix}")
    shutil.copy2(src, dest)
    logger.info("Backup created: %s", dest.name)
    return dest


def load_template(path: Path) -> TemplateInfo:
    logger.info("Loading template...")
    try:
        workbook = load_workbook(path)
    except Exception as exc:
        raise ExcelTemplateError(f"Failed to open XLSX (file may be damaged): {exc}") from exc

    worksheet = workbook.active
    headers, header_names = _read_headers(worksheet)
    missing = [name for name in REQUIRED_COLUMNS if name not in headers]
    if missing:
        raise ExcelTemplateError(
            "Required columns were not found in row 1: " + ", ".join(missing)
        )

    data_rows = _find_data_rows(worksheet)
    logger.info("Found %s columns", worksheet.max_column)
    logger.info("Found %s data row(s)", len(data_rows))
    return TemplateInfo(
        path=path,
        workbook=workbook,
        worksheet=worksheet,
        headers=headers,
        header_names=header_names,
        max_row=worksheet.max_row,
        max_column=worksheet.max_column,
        dimensions=worksheet.dimensions,
        data_rows=data_rows,
    )


def read_row_source(
    info: TemplateInfo,
    row: int,
    title_override: str | None = None,
    body_override: str | None = None,
    link_override: str | None = None,
    geo_override: str | None = None,
    product_override: str | None = None,
) -> RowSource:
    excel_title = _cell_str(info, row, "Title")
    excel_body = _cell_str(info, row, "Body")
    excel_link = _cell_str(info, row, "Link")
    title = title_override if title_override is not None else excel_title
    body = body_override if body_override is not None else excel_body
    link = link_override if link_override is not None else excel_link
    geo = geo_override
    if geo is None and "Countries" in info.headers:
        geo = _cell_str(info, row, "Countries") or None
    return RowSource(
        row=row,
        title=title,
        body=body,
        link=link,
        geo=geo,
        product=(product_override or "").strip(),
    )


def write_localizations(
    info: TemplateInfo,
    row: int,
    languages: list[str],
    localizations: list[Localization],
) -> None:
    if len(languages) != len(localizations):
        raise ExcelTemplateError("Language count does not match localization count")

    for slot in range(MAX_LANGUAGES):
        if slot < len(languages):
            item = localizations[slot]
            _set(info, row, LANGUAGE_COLUMNS[slot], item.language)
            _set(info, row, TITLE_COLUMNS[slot], item.title)
            _set(info, row, BODY_COLUMNS[slot], item.body)
            _set(info, row, LINK_COLUMNS[slot], item.link)
            if DISPLAY_LINK_COLUMNS[slot] in info.headers:
                _set(info, row, DISPLAY_LINK_COLUMNS[slot], _display_link(item.link))
        else:
            _set(info, row, LANGUAGE_COLUMNS[slot], None)
            _set(info, row, TITLE_COLUMNS[slot], None)
            _set(info, row, BODY_COLUMNS[slot], None)
            _set(info, row, LINK_COLUMNS[slot], None)
            if DISPLAY_LINK_COLUMNS[slot] in info.headers:
                _set(info, row, DISPLAY_LINK_COLUMNS[slot], None)


def normalize_region_code(region: str) -> str:
    value = region.strip().upper()
    if len(value) != 2 or not value.isalpha():
        raise ExcelTemplateError(
            f"ISO country code must be 2 letters, got {region!r}"
        )
    return value


def write_region(info: TemplateInfo, row: int, region_code: str) -> None:
    if SPECIAL_AD_CATEGORY_COUNTRY not in info.headers:
        raise ExcelTemplateError(
            f"Column {SPECIAL_AD_CATEGORY_COUNTRY!r} was not found in the template"
        )
    _set(info, row, SPECIAL_AD_CATEGORY_COUNTRY, region_code)
    logger.info(
        "Row %s %s = %s",
        row,
        SPECIAL_AD_CATEGORY_COUNTRY,
        region_code,
    )


def save_workbook(info: TemplateInfo, dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    info.workbook.save(dest)
    logger.info("Saving result...")
    logger.info("Done: %s", dest.name)


def snapshot_allowed_and_rest(path: Path) -> tuple[list[str], int, int, dict[tuple[int, int], Any]]:
    workbook = load_workbook(path, data_only=False)
    worksheet = workbook.active
    headers, header_names = _read_headers(worksheet)
    values: dict[tuple[int, int], Any] = {}
    for row in worksheet.iter_rows(
        min_row=1,
        max_row=worksheet.max_row,
        min_col=1,
        max_col=worksheet.max_column,
    ):
        for cell in row:
            values[(cell.row, cell.column)] = cell.value
    return header_names, worksheet.max_row, worksheet.max_column, values


def _read_headers(worksheet: Worksheet) -> tuple[dict[str, int], list[str]]:
    headers: dict[str, int] = {}
    names: list[str] = []
    for cell in worksheet[1]:
        raw = cell.value
        name = "" if raw is None else str(raw).strip()
        names.append(name)
        if not name:
            continue
        if name in headers:
            raise ExcelTemplateError(f"Duplicate column name in header row: {name}")
        headers[name] = cell.column
    if not any(names):
        raise ExcelTemplateError("Header row is empty")
    return headers, names


def _find_data_rows(worksheet: Worksheet) -> list[int]:
    rows: list[int] = []
    for row_idx in range(2, worksheet.max_row + 1):
        if _row_has_values(worksheet, row_idx, worksheet.max_column):
            rows.append(row_idx)
    return rows


def _row_has_values(worksheet: Worksheet, row: int, max_column: int) -> bool:
    for col in range(1, max_column + 1):
        value = worksheet.cell(row, col).value
        if value is not None and str(value).strip() != "":
            return True
    return False


def _cell_str(info: TemplateInfo, row: int, column_name: str) -> str:
    col = info.headers[column_name]
    value = info.worksheet.cell(row, col).value
    if value is None:
        return ""
    return str(value).strip()


def _set(info: TemplateInfo, row: int, column_name: str, value: Any) -> None:
    col = info.headers[column_name]
    info.worksheet.cell(row, col).value = value


def _display_link(url: str) -> str:
    parsed = urlparse(url.strip())
    host = parsed.netloc or parsed.path
    if host.startswith("www."):
        host = host[4:]
    return host
